"""
从 jplt/ 目录下的 Excel 文件生成 TypeScript 词汇数据文件
支持 N5-N1 所有级别
用法: python3 generateVocabFromExcel.py <级别> [起始行] [结束行]
示例: python3 generateVocabFromExcel.py N4        # 生成全部 N4
      python3 generateVocabFromExcel.py N4 1 500  # 生成 N4 第1-500行
"""

import openpyxl
import json
import re
import sys
import os

# 手动 romaji 映射（处理 pykakasi 无法处理的特殊情况）
MANUAL_ROMAJI = {
    'アイスクリーム': 'aisukurīmu',
    'アクセサリー': 'akusesarī',
    'アパート': 'apāto',
    'アルバイト': 'arubaito',
    'エレベーター': 'erebētā',
    'オートバイ': 'ōtobai',
    'カーテン': 'kāten',
    'カード': 'kādo',
    'カレーライス': 'karēraisu',
    'ガソリンスタンド': 'gasorinsutando',
    'ガラス': 'garasu',
    'ギター': 'gitā',
    'クラス': 'kurasu',
    'グラム': 'guramu',
    'クリスマス': 'kurisumasu',
    'グループ': 'gurūpu',
    'ゴミ': 'gomi',
    'コーヒー': 'kōhī',
    'コート': 'kōto',
    'コピー': 'kopī',
    'コンビニ': 'konbini',
    'コンピュータ': 'konpyūta',
    'サッカー': 'sakkā',
    'サラダ': 'sarada',
    'サンダル': 'sandaru',
    'シャツ': 'shatsu',
    'シャワー': 'shawā',
    'ジュース': 'jūsu',
    'スイッチ': 'suitchi',
    'スカート': 'sukāto',
    'スキー': 'sukī',
    'スクリーン': 'sukurīn',
    'スプーン': 'supūn',
    'スポーツ': 'supōtsu',
    'ズボン': 'zubon',
    'セーター': 'sētā',
    'ソファー': 'sofā',
    'タオル': 'taoru',
    'タクシー': 'takushī',
    'タバコ': 'tabako',
    'タマネギ': 'tamanegi',
    'ダンス': 'dansu',
    'チーズ': 'chīzu',
    'チケット': 'chiketto',
    'チャーハン': 'chāhan',
    'テープ': 'tēpu',
    'テーブル': 'tēburu',
    'テープレコーダー': 'tēpurekōdā',
    'テニス': 'tenisu',
    'テレビ': 'terebi',
    'トイレ': 'toire',
    'トマト': 'tomato',
    'ドア': 'doa',
    'ドライブ': 'doraibu',
    'ナイフ': 'naifu',
    'ニュース': 'nyūsu',
    'ネクタイ': 'nekutai',
    'ノート': 'nōto',
    'パーティー': 'pātī',
    'パソコン': 'pasokon',
    'バス': 'basu',
    'バター': 'batā',
    'ハンカチ': 'hankachi',
    'ハンサム': 'hansamu',
    'バナナ': 'banana',
    'ビール': 'bīru',
    'ピアノ': 'piano',
    'ビデオ': 'bideo',
    'フォーク': 'fōku',
    'プール': 'pūru',
    'プラスチック': 'purasuchikku',
    'プレゼント': 'purezento',
    'ベッド': 'beddo',
    'ペット': 'petto',
    'ペン': 'pen',
    'ボールペン': 'bōrupen',
    'ボタン': 'botan',
    'ポケット': 'poketto',
    'ホテル': 'hoteru',
    'ボランティア': 'borantia',
    'マッチ': 'matchi',
    'マンション': 'manshon',
    'ミルク': 'miruku',
    'メートル': 'mētoru',
    'メモ': 'memo',
    'メンバー': 'menbā',
    'モデル': 'moderu',
    'ラジオ': 'rajio',
    'ランチ': 'ranchi',
    'レストラン': 'resutoran',
    'レポート': 'repōto',
    'ワイン': 'wain',
    'ワイシャツ': 'waishatsu',
    'ワンピース': 'wanpīsu',
    'インク': 'inku',
    'エネルギー': 'enerugī',
    'カメラ': 'kamera',
    'カレンダー': 'karendā',
    'キロ': 'kiro',
    'キログラム': 'kiroguramu',
    'キロメートル': 'kiromētoru',
    'クラブ': 'kurabu',
    'ケーキ': 'kēki',
    'コップ': 'koppu',
    'サイズ': 'saizu',
    'サイン': 'sain',
    'シーズン': 'shīzun',
    'ジャム': 'jamu',
    'ストーブ': 'sutōbu',
    'スピーチ': 'supīchi',
    'スリッパ': 'surippa',
    'センチ': 'senchi',
    'ソース': 'sōsu',
    'タイプ': 'taipu',
    'ダイヤル': 'daiyaru',
    'デパート': 'depāto',
    'デザート': 'dezāto',
    'テント': 'tento',
    'ドラマ': 'dorama',
    'ハイキング': 'haikingu',
    'パジャマ': 'pajama',
    'パン': 'pan',
    'ハンドバッグ': 'handobaggu',
    'ハンドル': 'handoru',
    'ピクニック': 'pikunikku',
    'フィルム': 'firumu',
    'フライパン': 'furaipan',
    'ブラシ': 'burashi',
    'フランスパン': 'furansupan',
    'ブローチ': 'burōchi',
    'ペンチ': 'penchi',
    'ボール': 'bōru',
    'ポスト': 'posuto',
    'マスク': 'masuku',
    'マフラー': 'mafurā',
    'ミシン': 'mishin',
    'メニュー': 'menyū',
    'ヨット': 'yotto',
    'ライター': 'raitā',
    'ラケット': 'raketto',
    'リットル': 'rittoru',
    'レコード': 'rekōdo',
    'レジ': 'reji',
    'レモン': 'remon',
    'ロビー': 'robī',
    'アルコール': 'arukōru',
    'アルバム': 'arubamu',
    'アンケート': 'ankēto',
    'イコール': 'ikōru',
    'インスタント': 'insutanto',
    'ウイスキー': 'uisukī',
    'エアコン': 'eakon',
    'エスカレーター': 'esukarētā',
    'エンジン': 'enjin',
    'オーバー': 'ōbā',
    'オイル': 'oiru',
    'ガイド': 'gaido',
    'ガス': 'gasu',
    'ガレージ': 'garēji',
    'キス': 'kisu',
    'キャプテン': 'kyaputen',
    'キャンパス': 'kyanpasu',
    'キャンプ': 'kyanpu',
    'クーラー': 'kūrā',
    'クリーム': 'kurīmu',
    'グリーン': 'gurīn',
    'ゴール': 'gōru',
    'ゴールデンウィーク': 'gōrudenwīku',
    'コーラス': 'kōrasu',
    'コース': 'kōsu',
    'コーチ': 'kōchi',
    'コーヒーカップ': 'kōhīkappu',
    'コイン': 'koin',
    'コンサート': 'konsāto',
    'コンタクト': 'kontakuto',
    'サイクリング': 'saikuringu',
    'サラリーマン': 'sararīman',
    'シーツ': 'shītsu',
    'ジーパン': 'jīpan',
    'シート': 'shīto',
    'ジョギング': 'jogingu',
    'シルク': 'shiruku',
    'シングル': 'shinguru',
    'スーツ': 'sūtsu',
    'スーツケース': 'sūtsukēsu',
    'スーパー': 'sūpā',
    'スキル': 'sukiru',
    'スケート': 'sukēto',
    'スケジュール': 'sukejūru',
    'スタート': 'sutāto',
    'スタイル': 'sutairu',
    'ステーキ': 'sutēki',
    'ステレオ': 'sutereo',
    'ストップ': 'sutoppu',
    'ストレス': 'sutoresu',
    'スピーカー': 'supīkā',
    'スピード': 'supīdo',
    'スマート': 'sumāto',
    'スムーズ': 'sumūzu',
    'スローガン': 'surōgan',
    'セット': 'setto',
    'ゼロ': 'zero',
    'センター': 'sentā',
    'ソフト': 'sofuto',
    'タイトル': 'taitoru',
    'タイム': 'taimu',
    'ダブル': 'daburu',
    'チェック': 'chekku',
    'チャンス': 'chansu',
    'チャンネル': 'channeru',
    'ティッシュペーパー': 'tisshupepā',
    'テキスト': 'tekisuto',
    'テスト': 'tesuto',
    'テニスコート': 'tenisukōto',
    'テレビゲーム': 'terebigēmu',
    'トースト': 'tōsuto',
    'トップ': 'toppu',
    'トラック': 'torakku',
    'トランプ': 'toranpu',
    'ドレス': 'doresu',
    'トレーニング': 'torēningu',
    'ナイロン': 'nairon',
    'ナンバー': 'nanbā',
    'ネックレス': 'nekkuresu',
    'パーセント': 'pāsento',
    'バイク': 'baiku',
    'パイプ': 'paipu',
    'パイロット': 'pairotto',
    'バケツ': 'baketsu',
    'バスケットボール': 'basukettobōru',
    'バッグ': 'baggu',
    'パンフレット': 'panfuretto',
    'ビザ': 'biza',
    'ビタミン': 'bitamin',
    'ビル': 'biru',
    'ファックス': 'fakkusu',
    'ファン': 'fan',
    'ブーツ': 'būtsu',
    'フェリー': 'ferī',
    'ブラウス': 'burausu',
    'プラットホーム': 'purattohōmu',
    'プラン': 'puran',
    'フリー': 'furī',
    'プリント': 'purinto',
    'プレー': 'purē',
    'プログラム': 'puroguramu',
    'プロジェクト': 'purojekuto',
    'ベース': 'bēsu',
    'ベスト': 'besuto',
    'ベル': 'beru',
    'ベルト': 'beruto',
    'ペンキ': 'penki',
    'ボート': 'bōto',
    'ボーナス': 'bōnasu',
    'ホーム': 'hōmu',
    'ホームステイ': 'hōmusutei',
    'ポスター': 'posutā',
    'ポット': 'potto',
    'マーケット': 'māketto',
    'マイク': 'maiku',
    'マスター': 'masutā',
    'マナー': 'manā',
    'ミュージック': 'myūjikku',
    'メール': 'mēru',
    'モーター': 'mōtā',
    'ユニーク': 'yunīku',
    'ユニフォーム': 'yunifōmu',
    'ヨーロッパ': 'yōroppa',
    'ラーメン': 'rāmen',
    'ライス': 'raisu',
    'ライト': 'raito',
    'ラッシュアワー': 'rasshuawā',
    'リーダー': 'rīdā',
    'リズム': 'rizumu',
    'リボン': 'ribon',
    'ルーム': 'rūmu',
    'レインコート': 'reinkōto',
    'レクリエーション': 'rekuriēshon',
    'レンズ': 'renzu',
    'ロープ': 'rōpu',
    'ロケット': 'roketto',
    'ワット': 'watto',
}

# 各级别的文件配置
LEVEL_CONFIG = {
    'N5': {'file': 'jplt/N5词汇885个.xlsx', 'count': 885, 'var': 'n5Vocabulary'},
    'N4': {'file': 'jplt/N4词汇1917个.xlsx', 'count': 1917, 'var': 'n4Vocabulary'},
    'N3': {'file': 'jplt/N3词汇4170个.xlsx', 'count': 4170, 'var': 'n3Vocabulary'},
    'N2': {'file': 'jplt/N2词汇6702个.xlsx', 'count': 6702, 'var': 'n2Vocabulary'},
    'N1': {'file': 'jplt/N1词汇10000个.xlsx', 'count': 10000, 'var': 'n1Vocabulary'},
}


def to_romaji(text):
    """将假名转换为 romaji"""
    if text in MANUAL_ROMAJI:
        return MANUAL_ROMAJI[text]
    
    try:
        import pykakasi
        kks = pykakasi.kakasi()
        result = kks.convert(text)
        romaji = ''
        for item in result:
            romaji += item['hepburn']
        return romaji
    except:
        return text


def is_kana(text):
    """检查字符串是否包含日文假名"""
    return bool(re.search(r'[\u3040-\u309f\u30a0-\u30ff]', text))


def escape_ts_string(s):
    """转义 TypeScript 字符串中的特殊字符"""
    if s is None:
        return ''
    s = str(s)
    s = s.replace('\\', '\\\\')
    s = s.replace("'", "\\'")
    s = s.replace('\n', '\\n')
    s = s.replace('\r', '\\r')
    return s


def process_level(level, start_row=None, end_row=None):
    """处理指定级别的词汇数据"""
    if level not in LEVEL_CONFIG:
        print(f'错误: 不支持的级别 {level}')
        print(f'支持的级别: {", ".join(LEVEL_CONFIG.keys())}')
        return
    
    config = LEVEL_CONFIG[level]
    filepath = config['file']
    var_name = config['var']
    
    if not os.path.exists(filepath):
        print(f'错误: 文件不存在 {filepath}')
        return
    
    print(f'正在处理 {level} 词汇...')
    print(f'文件: {filepath}')
    
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    # 检查列数（N1 文件有 7 列，多一个"等级"列）
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    num_cols = len(header)
    has_level_col = (num_cols == 7)
    
    if has_level_col:
        print(f'  检测到 7 列格式（含等级列），将按等级列筛选 {level} 词汇')
    
    items = []
    total_rows = ws.max_row - 1  # 减去表头
    
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        row_num = i + 2  # Excel 实际行号
        
        # 如果指定了范围，跳过范围外的行
        if start_row is not None and row_num < start_row:
            continue
        if end_row is not None and row_num > end_row:
            continue
        
        if has_level_col:
            # 7 列格式: kanji, hiragana, pitch, pos, meaning, example, row_level
            kanji, hiragana, pitch, pos, meaning, example, row_level = row
            # 如果指定了等级列，只处理匹配等级的词汇
            if row_level and str(row_level).strip() != level:
                continue
        else:
            # 6 列格式: kanji, hiragana, pitch, pos, meaning, example
            kanji, hiragana, pitch, pos, meaning, example = row
        
        if kanji is None or str(kanji).strip() == '':
            continue
        
        kanji = str(kanji).strip()
        hiragana = str(hiragana).strip() if hiragana else ''
        meaning = str(meaning).strip() if meaning else ''
        example = str(example).strip() if example else ''
        
        # 如果假名读音列不是假名（如片假名词汇的读音列可能为空或错误），则使用 kanji 列
        if hiragana and not is_kana(hiragana):
            # 对于片假名词汇，kanji 本身就是片假名读音
            hiragana = kanji
        
        # 生成 romaji
        romaji = to_romaji(hiragana)
        
        # 处理例句
        chinese_example = ''
        jp_example = ''
        if example:
            if '／' in example:
                parts = example.split('／')
                jp_example = parts[0].replace('［例］', '').replace('～', kanji).strip()
                chinese_example = parts[1].strip() if len(parts) > 1 else ''
            else:
                jp_example = example.replace('［例］', '').replace('～', kanji).strip()
        
        pitch = str(pitch).strip() if pitch else ''
        pos = str(pos).strip() if pos else ''
        
        items.append({
            'kanji': kanji,
            'hiragana': hiragana,
            'romaji': romaji,
            'meaning': meaning,
            'level': level,
            'pitch': pitch,
            'pos': pos,
            'example': jp_example,
            'chineseExample': chinese_example,
        })
        
        # 每处理 500 行打印一次进度
        if len(items) % 500 == 0:
            print(f'  已处理 {len(items)} 条...')
    
    # 生成 TypeScript 文件
    lines = []
    lines.append(f'// 自动生成自 {os.path.basename(filepath)}')
    lines.append('// 请勿手动修改')
    lines.append('')
    lines.append("import { VocabularyItem } from '../mockData';")
    lines.append('')
    lines.append(f'export const {var_name}: VocabularyItem[] = [')
    
    for item in items:
        kanji_escaped = escape_ts_string(item['kanji'])
        hiragana_escaped = escape_ts_string(item['hiragana'])
        romaji_escaped = escape_ts_string(item['romaji'])
        meaning_escaped = escape_ts_string(item['meaning'])
        example_escaped = escape_ts_string(item['example'])
        chinese_example_escaped = escape_ts_string(item['chineseExample'])
        
        lines.append('  {')
        lines.append(f"    kanji: '{kanji_escaped}',")
        lines.append(f"    hiragana: '{hiragana_escaped}',")
        lines.append(f"    romaji: '{romaji_escaped}',")
        lines.append(f"    meaning: '{meaning_escaped}',")
        lines.append(f"    level: '{level}',")
        if item['pitch']:
            lines.append(f"    pitch: '{escape_ts_string(item['pitch'])}',")
        if item['pos']:
            lines.append(f"    pos: '{escape_ts_string(item['pos'])}',")
        if example_escaped:
            lines.append(f"    example: '{example_escaped}',")
        if chinese_example_escaped:
            lines.append(f"    chineseExample: '{chinese_example_escaped}',")
        lines.append('  },')
    
    lines.append('];')
    lines.append('')
    
    output = '\n'.join(lines)
    
    output_file = f'backend/src/data/generated/vocab-{level.lower()}.ts'
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(output)
    
    print(f'\n✅ 成功生成 {len(items)} 条 {level} 词汇数据')
    print(f'📄 输出文件: {output_file}')
    
    return len(items)


def update_index():
    """更新 vocab-index.ts 中的总词汇数"""
    total = 0
    for level in ['N5', 'N4', 'N3', 'N2', 'N1']:
        var_name = LEVEL_CONFIG[level]['var']
        filepath = f'backend/src/data/generated/vocab-{level.lower()}.ts'
        if os.path.exists(filepath):
            with open(filepath, 'r', encoding='utf-8') as f:
                content = f.read()
                # 统计词汇数量：计算 "level: 'Nx'" 的出现次数
                count = content.count(f"level: '{level}'")
                total += count
                print(f'  {level}: {count} 条')
    
    # 更新 index 文件中的总数
    index_file = 'backend/src/data/generated/vocab-index.ts'
    if os.path.exists(index_file):
        with open(index_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        import re
        content = re.sub(
            r'export const totalVocabCount = \d+;',
            f'export const totalVocabCount = {total};',
            content
        )
        
        with open(index_file, 'w', encoding='utf-8') as f:
            f.write(content)
        
        print(f'\n📊 总词汇数已更新为: {total}')


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('用法: python3 generateVocabFromExcel.py <级别> [起始行] [结束行]')
        print('示例:')
        print('  python3 generateVocabFromExcel.py N4        # 生成全部 N4')
        print('  python3 generateVocabFromExcel.py N4 1 500  # 生成 N4 第1-500行')
        print('  python3 generateVocabFromExcel.py N4 501    # 生成 N4 从第501行开始')
        sys.exit(1)
    
    level = sys.argv[1].upper()
    start_row = int(sys.argv[2]) if len(sys.argv) > 2 else None
    end_row = int(sys.argv[3]) if len(sys.argv) > 3 else None
    
    process_level(level, start_row, end_row)
